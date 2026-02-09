import re
import json
import math
import argparse
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Tuple, Any

import pandas as pd

# Optional dependencies (script will degrade gracefully)
try:
    import requests
except Exception:
    requests = None

try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    import camelot
except Exception:
    camelot = None

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


# -----------------------------
# Configuration: keywords / regex
# -----------------------------

SECTION_HEADER_HINTS = [
    "genotyping and imputation",
    "imputation",
    "genotype",
    "methods",
    "materials and methods",
    "results",
    "study design",
    "statistical analysis",
    "analysis",
]

IMPUTATION_KEYWORDS = [
    # Expand beyond "impute/panel/1000/HRC/TOPMed" to reduce false NR
    "imput", "imputation", "imputed", "reference panel", "panel",
    "1000 genomes", "1000g", "phase 3", "p3", "HRC", "haplotype reference consortium",
    "topmed", "minimac", "eagle", "shapeit", "beagle", "michigan imputation server",
    "genotyping and imputation"
]

STAGE_KEYWORDS = {
    "meta-analysis": ["meta-analysis", "meta analysis", "inverse-variance", "fixed effect", "random effect", "metal"],
    "joint-analysis": ["joint analysis", "pooled", "combined genotype", "individual-level", "genotype data"],
    "replication": ["replication", "validate", "validation", "confirm", "follow-up", "follow up", "independent cohort"],
    "discovery": ["discovery", "genome-wide", "gwas", "screen", "scan", "initial"],
    # Some papers use "stage 1/2"
    "stage 1": ["stage 1", "stage i"],
    "stage 2": ["stage 2", "stage ii"],
}

ASSOCIATION_TYPE_RULES = [
    # order matters: first match wins
    ("Imaging", ["mri", "imaging", "ct", "pet", "dti", "wmh", "pvs", "hippocamp", "ventricle"]),
    ("Cognitive", ["cognitive", "memory", "mmse", "executive", "attention", "language", "processing speed"]),
    ("Fluid biomarker", ["csf", "plasma", "serum", "biomarker", "abeta", "tau", "ptau", "nfl"]),
    ("Neuropathology", ["neuropath", "braak", "cerad", "plaques", "tangles", "autopsy"]),
    ("Expression", ["eqtl", "expression", "transcript", "rna-seq", "bulk rna", "single cell"]),
    ("AD", ["alzheimer", "load", "ad case", "case-control", "disease risk"]),
    ("Other", []),
]

MODEL_HINTS = {
    "logistic": ["logistic regression", "logistic"],
    "linear": ["linear regression", "linear model"],
    "cox": ["cox", "proportional hazards", "hazard ratio", "survival"],
    "mixed": ["mixed model", "lmm", "gmmat", "saige", "bgenie"],
}

COVARIATE_HINTS = [
    "adjust", "adjusted", "covariate", "age", "sex", "site", "batch", "pcs", "principal components",
    "ancestry", "education", "scanner"
]


# -----------------------------
# Schema: columns for curated sheet
# (You can extend / reorder as needed)
# -----------------------------

CURATED_COLUMNS = [
    "Name", "RecordID", "PaperIDX", "TableIDX", "Notes",
    "Stage_original", "Stage",
    "Analyses type", "Model type", "Meta/Joint",
    "SNP-based, Gene-based",
    "Cohort", "Cohort_simplified (no counts)",
    "Sample size", "Cases", "Controls", "Sample information",
    "Imputation_simple2",
    "Population", "Population_map",
    "Analysis group",
    "Phenotype", "Phenotype-derived",
    "For plotting Beta and OR - derived",
    "Reported gene (gene based test)",
    "TopSNP", "Interactions",
    "Chr", "P-value", "BP(Position)",
    "RA 1(Reported Allele 1)", "RA 2(Reported Allele 2)",
    "Note on alleles and AF",
    "ReportedAF(MAF)", "AFincases", "AFincontrols",
    "Effect Size Type (OR or Beta)", "EffectSize(altvsref)",
    "95%ConfidenceInterval",
    "Genome build (hg18/hg37/hg38)",
    "Platform",
    "Approved symbol",
    "Pubmed PMID", "PMCID",
    "Table Ref in paper", "Table links",
    "LocusName",
    # additional helper flags
    "_confidence", "_needs_review", "_evidence"
]


# -----------------------------
# Audit structure
# -----------------------------
@dataclass
class FieldAudit:
    field: str
    value: Any
    confidence: float
    evidence: str
    rule: str
    needs_review: bool


# -----------------------------
# Utilities
# -----------------------------
def download_pdf(url: str, out_path: str) -> str:
    if requests is None:
        raise RuntimeError("requests is not installed; cannot download from URL.")
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    with open(out_path, "wb") as f:
        f.write(r.content)
    return out_path


def normalize_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


def find_snippets(text: str, keywords: List[str], window: int = 240) -> List[str]:
    """Return short evidence snippets around keyword hits."""
    hits = []
    t = text
    t_low = t.lower()
    for kw in keywords:
        kw_low = kw.lower()
        for m in re.finditer(re.escape(kw_low), t_low):
            start = max(0, m.start() - window)
            end = min(len(t), m.end() + window)
            hits.append(t[start:end].replace("\n", " "))
    # deduplicate
    uniq = []
    seen = set()
    for h in hits:
        k = h[:200]
        if k not in seen:
            uniq.append(h)
            seen.add(k)
    return uniq[:8]


def safe_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        s = s.replace("×", "x").replace("−", "-")
        # parse scientific like 4.2 x 10^-5
        sci = re.search(r"([0-9.]+)\s*x\s*10\^?(-?\d+)", s, flags=re.I)
        if sci:
            base = float(sci.group(1))
            exp = int(sci.group(2))
            return base * (10 ** exp)
        return float(s)
    except Exception:
        return None


# -----------------------------
# Core extraction: PDF text
# -----------------------------
def extract_full_text(pdf_path: str) -> str:
    if pdfplumber is None:
        return ""
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            parts.append(txt)
    return "\n".join(parts)


# -----------------------------
# Core extraction: tables
# -----------------------------
def extract_tables(pdf_path: str, pages: str = "all") -> List[pd.DataFrame]:
    """
    Try Camelot lattice then stream. Fall back to empty list if Camelot is unavailable.
    """
    if camelot is None:
        return []

    tables = []
    # Try lattice (best for ruling lines)
    try:
        t_lattice = camelot.read_pdf(pdf_path, pages=pages, flavor="lattice")
        for t in t_lattice:
            df = t.df
            tables.append(df)
    except Exception:
        pass

    # Try stream (best for whitespace-separated)
    if not tables:
        try:
            t_stream = camelot.read_pdf(pdf_path, pages=pages, flavor="stream")
            for t in t_stream:
                df = t.df
                tables.append(df)
        except Exception:
            pass

    return tables


# -----------------------------
# Rules / inference
# -----------------------------
def infer_imputation(full_text: str) -> Tuple[str, FieldAudit]:
    t = normalize_text(full_text)
    snippets = find_snippets(full_text, IMPUTATION_KEYWORDS, window=220)

    # Heuristic: detect panels
    panels = []
    if re.search(r"\btopmed\b", t):
        panels.append("TOPMed")
    if re.search(r"\bhrc\b|\bhaplotype reference consortium\b", t):
        panels.append("HRC")
    if re.search(r"\b1000g\b|\b1000 genomes\b|\b1000genomes\b", t):
        panels.append("1000G")

    if panels:
        value = ";".join(sorted(set(panels)))
        conf = 0.85
        rule = "panel keyword detection (TOPMed/HRC/1000G)"
        needs_review = False
        evidence = snippets[0] if snippets else "Detected panel keywords in text."
        # If multiple panels mentioned, often table-specific ambiguity
        if len(set(panels)) > 1:
            conf = 0.65
            needs_review = True
            rule += " (multiple panels; table-specific mapping unclear)"
        return value, FieldAudit("Imputation_simple2", value, conf, evidence, rule, needs_review)

    value = "NR"
    return value, FieldAudit("Imputation_simple2", value, 0.3, (snippets[0] if snippets else ""), "no imputation keywords found", True)


def infer_stage(full_text: str) -> Tuple[str, FieldAudit]:
    t = normalize_text(full_text)
    hits = []
    for stage, kws in STAGE_KEYWORDS.items():
        for kw in kws:
            if kw in t:
                hits.append(stage)
                break

    # Priority logic
    if "meta-analysis" in hits:
        value = "Meta-analysis"
        conf = 0.8
        needs_review = False
        rule = "found meta-analysis cues"
    elif "joint-analysis" in hits:
        value = "Joint-analysis"
        conf = 0.75
        needs_review = True  # joint vs pooled wording can be ambiguous
        rule = "found joint-analysis cues"
    elif "stage 1" in hits or "stage 2" in hits:
        # if explicit stage labels exist, keep as-is
        # could be Stage 1/Stage 2 rather than discovery/replication
        value = "Stage 1/2"
        conf = 0.7
        needs_review = True
        rule = "explicit Stage 1/2 detected; mapping needs confirmation"
    elif "replication" in hits:
        value = "Replication/Validation"
        conf = 0.7
        needs_review = True
        rule = "found replication cues"
    elif "discovery" in hits:
        value = "Discovery"
        conf = 0.6
        needs_review = True
        rule = "found generic discovery/GWAS cues"
    else:
        value = "NR"
        conf = 0.25
        needs_review = True
        rule = "no stage cues found"

    snippets = find_snippets(full_text, ["meta-analysis", "replication", "validation", "stage 1", "stage 2", "gwas"], window=220)
    evidence = snippets[0] if snippets else ""
    return value, FieldAudit("Stage", value, conf, evidence, rule, needs_review)


def infer_association_type(full_text: str) -> Tuple[str, FieldAudit]:
    t = normalize_text(full_text)
    for label, kws in ASSOCIATION_TYPE_RULES:
        if not kws:
            continue
        for kw in kws:
            if kw in t:
                value = label
                conf = 0.7
                needs_review = True  # conservative
                rule = f"matched association keywords -> {label}"
                snippets = find_snippets(full_text, [kw], window=220)
                evidence = snippets[0] if snippets else ""
                return value, FieldAudit("Analyses type", value, conf, evidence, rule, needs_review)

    value = "NR"
    return value, FieldAudit("Analyses type", value, 0.25, "", "no association cues found", True)


def infer_model_type(full_text: str) -> Tuple[str, FieldAudit]:
    t = normalize_text(full_text)
    model_family = None
    for fam, kws in MODEL_HINTS.items():
        if any(kw in t for kw in kws):
            model_family = fam
            break

    has_cov = any(kw in t for kw in COVARIATE_HINTS)

    if model_family:
        if has_cov:
            value = f"{model_family} regression (with covariates)"
            conf = 0.75
            needs_review = True
            rule = "model family + covariate cues"
        else:
            value = f"{model_family} regression"
            conf = 0.65
            needs_review = True
            rule = "model family cues only"
    else:
        if has_cov:
            value = "model with covariates (family NR)"
            conf = 0.55
            needs_review = True
            rule = "covariate cues only"
        else:
            value = "NR"
            conf = 0.25
            needs_review = True
            rule = "no model cues"

    snippets = find_snippets(full_text, ["logistic", "linear regression", "cox", "hazard ratio", "adjusted", "covariate", "principal components"], window=220)
    evidence = snippets[0] if snippets else ""
    return value, FieldAudit("Model type", value, conf, evidence, rule, needs_review)


def classify_meta_joint(stage_value: str) -> str:
    if stage_value.lower().startswith("meta"):
        return "Meta"
    if stage_value.lower().startswith("joint"):
        return "Joint"
    return "NR"


# -----------------------------
# Table parsing to association records (heuristic)
# -----------------------------
RSID_RE = re.compile(r"\brs\d+\b", flags=re.I)

def guess_table_type(df: pd.DataFrame) -> str:
    header = " ".join(df.iloc[0].astype(str).tolist()).lower() if len(df) > 0 else ""
    if "hazard" in header or "hr" in header:
        return "survival"
    if "or" in header and "p" in header:
        return "or_table"
    if "beta" in header or "β" in header:
        return "beta_table"
    return "unknown"


def extract_records_from_table(df: pd.DataFrame, table_idx: int, paper_id: str, pmcid: str, table_ref: str, table_link: str) -> List[Dict[str, Any]]:
    """
    Minimal heuristic: find rsIDs and extract nearby columns if recognizable.
    This will NOT be perfect for every format; it is meant to pre-fill + flag.
    """
    records = []
    if df.empty:
        return records

    # Assume first row is header
    header = [str(x).strip() for x in df.iloc[0].tolist()]
    body = df.iloc[1:].copy()
    body.columns = header

    # Find likely columns
    cols = [c.lower() for c in body.columns]
    rs_col = None
    for c in body.columns:
        if "rs" in c.lower() or "snp" in c.lower():
            rs_col = c
            break

    # Fallback: scan rows for rsIDs
    for i, row in body.iterrows():
        row_text = " ".join([str(x) for x in row.tolist()])
        rsids = RSID_RE.findall(row_text)
        if not rsids:
            continue

        # Use first rsID as primary
        rsid = rsids[0].lower()

        # Attempt to locate P-value
        pval = None
        for c in body.columns:
            if "p" == c.lower() or "p-value" in c.lower() or "p value" in c.lower():
                pval = safe_float(row.get(c))
                break
        # Attempt OR / beta
        effect = None
        effect_type = "NR"
        ci = "NR"

        # OR
        for c in body.columns:
            if c.lower().strip() == "or" or "odds ratio" in c.lower():
                effect = safe_float(row.get(c))
                if effect is not None:
                    effect_type = "OR"
                break

        # Beta
        if effect is None:
            for c in body.columns:
                if "beta" in c.lower() or "β" in c.lower():
                    effect = safe_float(row.get(c))
                    if effect is not None:
                        effect_type = "Beta"
                    break

        # CI (if present)
        for c in body.columns:
            if "95" in c.lower() and "ci" in c.lower():
                ci = str(row.get(c)).strip()
                break

        # Chr / position (optional)
        chr_val = "NR"
        bp_val = "NR"
        for c in body.columns:
            if "chr" in c.lower():
                chr_val = str(row.get(c)).strip()
            if "pos" in c.lower() or "position" in c.lower() or "bp" in c.lower():
                bp_val = str(row.get(c)).strip()

        rec = {k: "NR" for k in CURATED_COLUMNS}
        rec["Name"] = rsid
        rec["PaperIDX"] = paper_id
        rec["PMCID"] = pmcid
        rec["TableIDX"] = f"T{table_idx:04d}"
        rec["Table Ref in paper"] = table_ref
        rec["Table links"] = table_link
        rec["P-value"] = pval if pval is not None else "NR"
        rec["Effect Size Type (OR or Beta)"] = effect_type
        rec["EffectSize(altvsref)"] = effect if effect is not None else "NR"
        rec["95%ConfidenceInterval"] = ci
        rec["Chr"] = chr_val
        rec["BP(Position)"] = bp_val
        rec["_confidence"] = 0.45
        rec["_needs_review"] = True
        rec["_evidence"] = f"Row text: {row_text[:260]}..."
        records.append(rec)

    return records


# -----------------------------
# Export to Excel
# -----------------------------
def load_template_headers(template_xlsx: Optional[str]) -> List[str]:
    if template_xlsx:
        wb = load_workbook(template_xlsx)
        ws = wb.active
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")
        headers = [h for h in headers if h]
        if headers:
            return headers
    return CURATED_COLUMNS


def write_curated_xlsx(records: List[Dict[str, Any]], out_path: str, template_xlsx: Optional[str] = None) -> None:
    headers = load_template_headers(template_xlsx)
    df = pd.DataFrame(records)
    # Ensure all headers exist
    for h in headers:
        if h not in df.columns:
            df[h] = "NR"
    df = df[headers]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="curated")


# -----------------------------
# Main pipeline
# -----------------------------
def run_pipeline(pdf_or_url: str,
                 out_xlsx: str,
                 audit_json: str,
                 paper_id: str = "PAPER",
                 pmcid: str = "NR",
                 template_xlsx: Optional[str] = None) -> None:

    # 1) Get PDF
    pdf_path = pdf_or_url
    if pdf_or_url.lower().startswith("http"):
        if requests is None:
            raise RuntimeError("URL provided but requests not installed.")
        pdf_path = "input_paper.pdf"
        download_pdf(pdf_or_url, pdf_path)

    # 2) Extract full text
    full_text = extract_full_text(pdf_path)

    # 3) Infer global fields (Stage / Model / Assoc type / Imputation)
    stage_val, stage_audit = infer_stage(full_text)
    assoc_val, assoc_audit = infer_association_type(full_text)
    model_val, model_audit = infer_model_type(full_text)
    imp_val, imp_audit = infer_imputation(full_text)
    meta_joint = classify_meta_joint(stage_val)

    # 4) Extract tables
    tables = extract_tables(pdf_path, pages="all")

    # 5) Convert each table to records (heuristic)
    records: List[Dict[str, Any]] = []
    audits: Dict[str, Any] = {
        "paper": {
            "paper_id": paper_id,
            "pmcid": pmcid,
            "pdf": pdf_path,
        },
        "global_field_audit": [
            asdict(stage_audit),
            asdict(assoc_audit),
            asdict(model_audit),
            asdict(imp_audit),
        ],
        "record_field_audit": []
    }

    for idx, df in enumerate(tables, start=1):
        table_ref = f"Table {idx}"
        table_link = pdf_or_url if pdf_or_url.lower().startswith("http") else "local_pdf"

        recs = extract_records_from_table(df, idx, paper_id, pmcid, table_ref, table_link)
        for r in recs:
            # Apply global inferred values as pre-fill
            r["Stage"] = stage_val
            r["Stage_original"] = stage_val
            r["Analyses type"] = assoc_val
            r["Model type"] = model_val
            r["Meta/Joint"] = meta_joint
            r["Imputation_simple2"] = imp_val

            # Conservative flags: if global inference uncertain, propagate review
            global_conf = min(stage_audit.confidence, assoc_audit.confidence, model_audit.confidence, imp_audit.confidence)
            r["_confidence"] = float(r.get("_confidence", 0.4)) * 0.5 + global_conf * 0.5
            if stage_audit.needs_review or assoc_audit.needs_review or model_audit.needs_review or imp_audit.needs_review:
                r["_needs_review"] = True

            # Add short evidence
            r["_evidence"] = (r.get("_evidence", "") + "\n" +
                             f"[Stage evidence] {stage_audit.evidence[:240]}\n" +
                             f"[Model evidence] {model_audit.evidence[:240]}\n" +
                             f"[Imputation evidence] {imp_audit.evidence[:240]}").strip()

            records.append(r)

    # 6) If no tables extracted, still output a shell row for manual work
    if not records:
        shell = {k: "NR" for k in CURATED_COLUMNS}
        shell["PaperIDX"] = paper_id
        shell["PMCID"] = pmcid
        shell["Stage"] = stage_val
        shell["Analyses type"] = assoc_val
        shell["Model type"] = model_val
        shell["Imputation_simple2"] = imp_val
        shell["_confidence"] = min(stage_audit.confidence, assoc_audit.confidence, model_audit.confidence, imp_audit.confidence)
        shell["_needs_review"] = True
        shell["_evidence"] = "No tables extracted. Use global evidence snippets to curate manually."
        records.append(shell)

    # 7) Write outputs
    write_curated_xlsx(records, out_xlsx, template_xlsx=template_xlsx)

    with open(audit_json, "w", encoding="utf-8") as f:
        json.dump(audits, f, indent=2, ensure_ascii=False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="PDF path or URL")
    parser.add_argument("--out", default="curated_output.xlsx")
    parser.add_argument("--audit", default="audit.json")
    parser.add_argument("--paper_id", default="PAPER")
    parser.add_argument("--pmcid", default="NR")
    parser.add_argument("--template", default=None, help="Optional template xlsx (first row as headers)")
    args = parser.parse_args()

    run_pipeline(
        pdf_or_url=args.input,
        out_xlsx=args.out,
        audit_json=args.audit,
        paper_id=args.paper_id,
        pmcid=args.pmcid,
        template_xlsx=args.template,
    )


if __name__ == "__main__":
    main()
